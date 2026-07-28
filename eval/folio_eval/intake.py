"""Firm-workbook intake for the F1 evaluation harness (U1, KTD1, KTD11, KTD12).

Original workbooks live outside the repo tree, in ``~/.folio-resolve-eval-data/`` (mode 700).
This module extracts *only* the in-scope taxonomy sheets named by the caller into gitignored
derived JSONL files under ``eval/data/<firm>/<sheet_name_hash>.jsonl`` — one row per line, each
row a JSON array of cell strings (``null`` for empty cells), header row first. It never reads
any other sheet's cell content; callers pass exactly the sheet names they intend to extract.

``eval/data/MANIFEST.md`` is committed and records, per firm: the external workbook path, the
workbook's SHA-256, and per in-scope sheet a SHA-256 of the sheet *name* (never the literal
name), a row count, a header signature, and a content hash of the derived JSONL — enough to
verify a derived file against the manifest without ever writing a firm surface string to a
committed file.

``openpyxl`` is imported lazily, function-local, only inside :func:`extract_sheet` (KTD11) —
every other function in this module, and every test outside the extraction path, runs under the
base venv against plain data structures.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

# eval/folio_eval/intake.py -> parent = eval/folio_eval, parent.parent = eval/
_EVAL_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_DATA_DIR = _EVAL_ROOT / "data"
DEFAULT_MANIFEST_PATH = DEFAULT_DATA_DIR / "MANIFEST.md"

CellValue = str | None
Row = list[CellValue]


class IntakeError(Exception):
    """Raised when a derived sheet fails verification against the manifest, or extraction fails."""


@dataclass(frozen=True, slots=True)
class SheetEntry:
    """One in-scope sheet's manifest record. Never carries the literal sheet name."""

    sheet_name_hash: str
    row_count: int
    header_signature: str
    content_sha256: str


@dataclass(frozen=True, slots=True)
class WorkbookEntry:
    """One firm's manifest record: workbook identity plus its in-scope sheets."""

    firm: str
    workbook_path: str
    workbook_sha256: str
    sheets: tuple[SheetEntry, ...]


# --------------------------------------------------------------------------------------
# Hashing primitives
# --------------------------------------------------------------------------------------


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_text(text: str) -> str:
    return sha256_bytes(text.encode("utf-8"))


def sha256_file(path: Path) -> str:
    return sha256_bytes(path.read_bytes())


def hash_sheet_name(name: str) -> str:
    """SHA-256 of the sheet name, UTF-8, no salt — the only form a sheet name may take on disk."""
    return sha256_text(name)


def header_signature(header: Sequence[CellValue]) -> str:
    """SHA-256 of the tab-joined first-row cells (``""`` for ``None``)."""
    joined = "\t".join(cell if cell is not None else "" for cell in header)
    return sha256_text(joined)


def rows_content_sha256(rows: Sequence[Sequence[CellValue]]) -> str:
    """SHA-256 over a deterministic JSON serialization of every row, in order."""
    text = "\n".join(json.dumps(list(row), ensure_ascii=False) for row in rows)
    return sha256_text(text)


def _cell_to_str(value: object) -> CellValue:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    return str(value)


def _atomic_write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(dir=path.parent, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(text)
        os.replace(tmp_name, path)
    finally:
        if os.path.exists(tmp_name):
            os.unlink(tmp_name)


# --------------------------------------------------------------------------------------
# Extraction (the only function-local openpyxl import in this package)
# --------------------------------------------------------------------------------------


def extract_sheet(workbook_path: Path, sheet_name: str, out_path: Path) -> SheetEntry:
    """Extract one in-scope sheet's rows verbatim into a derived JSONL file at ``out_path``.

    Reads only the named sheet — no other sheet's cell content is ever touched here, streaming
    via ``read_only=True`` so no other worksheet's cells are ever parsed into memory. Trailing
    all-blank header columns are trimmed to the sheet's actual used width; every row is clipped
    to that width and every cell coerced to ``str | None``. Some workbooks carry a stale
    ``<dimension>`` tag wider than the real data (formatting applied to a padded row range) —
    read-only streaming trusts that tag, so trailing rows that are entirely blank (within the
    trimmed width) are dropped after the fact, on this sheet's own already-read rows only.
    Writes atomically (tempfile + ``os.replace``, mirroring ``feedback_store.py``). Returns the
    :class:`SheetEntry` for the manifest — never the literal sheet name.
    """
    import openpyxl  # intentionally function-local (KTD11) -- the only openpyxl import in the repo

    name_hash = hash_sheet_name(sheet_name)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise IntakeError(f"sheet not found in workbook: sheet_name_hash={name_hash}")
        worksheet = workbook[sheet_name]
        raw_rows: list[tuple[object, ...]] = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not raw_rows:
        raise IntakeError(f"sheet has no rows: sheet_name_hash={name_hash}")

    header_raw = raw_rows[0]
    width = 0
    for idx, value in enumerate(header_raw):
        if value is not None and str(value).strip() != "":
            width = idx + 1
    if width == 0:
        raise IntakeError(f"sheet header is empty: sheet_name_hash={name_hash}")

    rows: list[Row] = [[_cell_to_str(v) for v in row[:width]] for row in raw_rows]

    # Drop trailing padding rows (stale wide <dimension> tag): keep the header even if blank.
    while len(rows) > 1 and all(cell is None for cell in rows[-1]):
        rows.pop()

    _atomic_write_text(out_path, "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n")

    return SheetEntry(
        sheet_name_hash=name_hash,
        row_count=len(rows),
        header_signature=header_signature(rows[0]),
        content_sha256=rows_content_sha256(rows),
    )


# --------------------------------------------------------------------------------------
# Manifest parse / write
# --------------------------------------------------------------------------------------

_FIRM_RE = re.compile(r"^## (.+)$")
_PATH_RE = re.compile(r"^- Workbook path: `(.+)`$")
_SHA_RE = re.compile(r"^- Workbook SHA-256: `([0-9a-f]{64})`$")
_ROW_RE = re.compile(
    r"^\| `([0-9a-f]{64})` \| (\d+) \| `([0-9a-f]{64})` \| `([0-9a-f]{64})` \|$"
)


def format_manifest(entries: Sequence[WorkbookEntry]) -> str:
    """Render manifest entries as the committed markdown table (KTD1: hashes only, no names)."""
    lines: list[str] = [
        "# eval/data/MANIFEST.md",
        "",
        "Derived-data manifest for the F1 evaluation harness (KTD1). Records only hashes, row",
        "counts, and header signatures for in-scope taxonomy sheets -- no literal sheet names,",
        "no out-of-scope sheets, no firm surface strings. Regenerated by",
        "`eval/folio_eval/intake.py`'s extraction routine; do not hand-edit.",
        "",
    ]
    for entry in entries:
        lines.append(f"## {entry.firm}")
        lines.append("")
        lines.append(f"- Workbook path: `{entry.workbook_path}`")
        lines.append(f"- Workbook SHA-256: `{entry.workbook_sha256}`")
        lines.append("")
        lines.append(
            "| Sheet name hash (sha256) | Row count | Header signature (sha256) | Content SHA-256 |"
        )
        lines.append("|---|---|---|---|")
        for sheet in entry.sheets:
            lines.append(
                f"| `{sheet.sheet_name_hash}` | {sheet.row_count} | "
                f"`{sheet.header_signature}` | `{sheet.content_sha256}` |"
            )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def parse_manifest(text: str) -> list[WorkbookEntry]:
    """Parse the committed markdown manifest back into structured records."""
    entries: list[WorkbookEntry] = []
    firm: str | None = None
    workbook_path: str | None = None
    workbook_sha256: str | None = None
    sheets: list[SheetEntry] = []

    def flush() -> None:
        nonlocal firm, workbook_path, workbook_sha256, sheets
        if firm is not None:
            if workbook_path is None or workbook_sha256 is None:
                raise IntakeError(f"manifest section for firm={firm!r} missing path or sha256")
            entries.append(
                WorkbookEntry(
                    firm=firm,
                    workbook_path=workbook_path,
                    workbook_sha256=workbook_sha256,
                    sheets=tuple(sheets),
                )
            )
        firm, workbook_path, workbook_sha256, sheets = None, None, None, []

    for line in text.splitlines():
        firm_match = _FIRM_RE.match(line)
        if firm_match:
            flush()
            firm = firm_match.group(1).strip()
            continue
        path_match = _PATH_RE.match(line)
        if path_match:
            workbook_path = path_match.group(1)
            continue
        sha_match = _SHA_RE.match(line)
        if sha_match:
            workbook_sha256 = sha_match.group(1)
            continue
        row_match = _ROW_RE.match(line)
        if row_match:
            sheets.append(
                SheetEntry(
                    sheet_name_hash=row_match.group(1),
                    row_count=int(row_match.group(2)),
                    header_signature=row_match.group(3),
                    content_sha256=row_match.group(4),
                )
            )
            continue
    flush()
    return entries


def write_manifest(path: Path, entries: Sequence[WorkbookEntry]) -> None:
    _atomic_write_text(path, format_manifest(entries))


def read_manifest(path: Path) -> list[WorkbookEntry]:
    if not path.exists():
        raise IntakeError(f"manifest not found: {path}")
    return parse_manifest(path.read_text(encoding="utf-8"))


# --------------------------------------------------------------------------------------
# Loader
# --------------------------------------------------------------------------------------


def load_sheet_rows(
    firm: str,
    sheet_name_hash: str,
    *,
    data_dir: Path = DEFAULT_DATA_DIR,
    manifest_path: Path = DEFAULT_MANIFEST_PATH,
) -> list[Row]:
    """Load a derived-sheet JSONL file, verifying it against the parsed manifest.

    Verifies, in order, the derived file's content hash, row count, and header signature
    against the firm's manifest entry for ``sheet_name_hash``; raises :class:`IntakeError`
    naming the failed check (and its expected/actual values) on any mismatch.
    """
    entries = read_manifest(manifest_path)
    workbook_entry = next((e for e in entries if e.firm == firm), None)
    if workbook_entry is None:
        raise IntakeError(f"manifest verification failed: check=firm firm={firm!r} not found")

    sheet_entry = next(
        (s for s in workbook_entry.sheets if s.sheet_name_hash == sheet_name_hash), None
    )
    if sheet_entry is None:
        raise IntakeError(
            "manifest verification failed: check=sheet_name_hash "
            f"firm={firm!r} sheet_name_hash={sheet_name_hash} not found"
        )

    jsonl_path = data_dir / firm / f"{sheet_name_hash}.jsonl"
    if not jsonl_path.exists():
        raise IntakeError(f"manifest verification failed: check=file_present path={jsonl_path}")

    lines = [line for line in jsonl_path.read_text(encoding="utf-8").splitlines() if line]
    rows: list[Row] = [json.loads(line) for line in lines]

    actual_content_sha256 = rows_content_sha256(rows)
    if actual_content_sha256 != sheet_entry.content_sha256:
        raise IntakeError(
            "manifest verification failed: check=content_sha256 "
            f"expected={sheet_entry.content_sha256} actual={actual_content_sha256}"
        )

    if len(rows) != sheet_entry.row_count:
        raise IntakeError(
            "manifest verification failed: check=row_count "
            f"expected={sheet_entry.row_count} actual={len(rows)}"
        )

    header = rows[0] if rows else []
    actual_header_signature = header_signature(header)
    if actual_header_signature != sheet_entry.header_signature:
        raise IntakeError(
            "manifest verification failed: check=header_signature "
            f"expected={sheet_entry.header_signature} actual={actual_header_signature}"
        )

    return rows
